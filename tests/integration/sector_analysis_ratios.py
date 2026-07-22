"""
Lameh Sector Analysis - Ratio Verification Script
==================================================

Purpose
-------
Verifies internal consistency of every "Financial Ratios" metric exported from
Lameh's Sector Analysis chart-builder ("select all ratios" export). The export
already includes, for every ratio, a full decomposition down to its input
components and raw financial-statement line items. This script recomputes each
ratio from its OWN reported input components (not from a separate ground truth)
and flags any case where reported_value != formula(reported_inputs).

This catches:
  - Aggregation/rollup bugs (e.g. a "Total X" that doesn't equal the sum of its
    parts, the Ma'aden COGS-footing type of issue).
  - Ratio-calculation bugs (wrong formula, wrong sign, wrong denominator).
  - Silent "N/A treated as 0" substitutions (flagged separately, not as errors,
    since this may be intentional but can be misleading - see NOTES below).

Usage
-----
    python3 verify_ratios.py path/to/export.xlsx [--tolerance 0.005] [--csv out.csv]

By default this runs two passes and (if --csv is given) writes two reports:
  - WEB:   whatever value is currently cached in the file (out-web.csv)
  - EXCEL: the value after opening the file in a real Excel instance and
           forcing a full recalculation (out-excel.csv)
The EXCEL pass requires pywin32 and a local Excel installation; skip it with
--web-only.

Input format expected
---------------------
An .xlsx with a "Chart Data" sheet laid out as:
    Row 3   : header row -> Metric, Entity, Section, Subsection, <period columns...>
    Row 4+  : one row per (metric, position-in-tree) instance, with period values
              in the columns following "Subsection".

NOTES on conventions discovered empirically from Lameh's own export data
-------------------------------------------------------------------------
  - Expense/outflow line items (Finance costs, Zakat, COGS, CapEx additions,
    Dividends paid, lease/loan repayments) are stored as NEGATIVE numbers.
    Ratios that measure a "cost coverage" or "turnover" typically take the
    absolute value of these; cash-flow SUM formulas (FCF, FCFE, FCFF, Total
    Debt Service) use the raw signed values directly (adding a negative value
    achieves the subtraction).
  - When an input is unavailable ('-' in the export), debt-related ratios
    (Debt to Assets, Debt to Capital, Debt to EBITDA, Debt to Equity, Net Debt,
    Net Debt to EBITDA) appear to silently treat the missing value as 0 rather
    than propagating "N/A". This script flags these cases separately as
    [SILENT-ZERO] rather than [FAIL], because it's not a math error, but it is
    worth a product decision: a company with "Total Debt: not retrieved" and
    a company with genuinely zero debt currently render identically (0.0x).
"""

import os
import sys
import argparse
import tempfile
from datetime import datetime
from openpyxl import load_workbook

TOLERANCE_DEFAULT = 0.005  # 0.5% relative tolerance
ABS_EPS = 1e-6              # absolute epsilon for near-zero comparisons
DEFAULT_CSV_DIR = os.path.join("results", "sector-analysis")


def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def get_company_name(path):
    """Read the company name from the export. Tries the "__lamehCompanyName"
    row in the Chart Data sheet first, then falls back to the Companies table
    in the "_lameh_context" sheet (used by some export layouts instead).
    Returns None if neither is present."""
    wb = load_workbook(path, data_only=True)

    if "Chart Data" in wb.sheetnames:
        ws = wb["Chart Data"]
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "__lamehCompanyName":
                return ws.cell(row=r, column=5).value

    if "_lameh_context" in wb.sheetnames:
        ws = wb["_lameh_context"]
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "companyId" and ws.cell(row=r, column=2).value == "companyName":
                return ws.cell(row=r + 1, column=2).value

    return None


def sanitize_filename(text):
    """Make arbitrary text (e.g. a company name) safe to use as a Windows filename."""
    import re
    return re.sub(r'[\\/:*?"<>|,\s]+', "_", str(text)).strip("_")


def load_metric_values(path):
    """Returns ({metric_name: [values per period column]}, periods, {metric_name: source_row})
    using first occurrence of each metric name (values are consistent across repeated
    occurrences in the same export, since they reference the same underlying figure).
    `source_row` is the 1-indexed Excel row the value was read from, for traceability."""
    wb = load_workbook(path, data_only=True)
    if "Chart Data" not in wb.sheetnames:
        raise ValueError('Expected a "Chart Data" sheet - is this a Lameh chart-builder export?')
    ws = wb["Chart Data"]

    header_row = 3
    first_data_row = 4
    max_col = ws.max_column
    periods = []
    for c in range(5, max_col + 1):
        val = ws.cell(row=header_row, column=c).value
        # Some export layouts append metadata columns (e.g. "__lamehCompanyId",
        # "__lamehCompanyName") right after the real period columns, with no
        # blank column in between - stop there too, not just at the first None.
        if val is None or (isinstance(val, str) and val.startswith("__lameh")):
            break
        periods.append(val)
    n_periods = len(periods)

    values = {}
    source_rows = {}
    for r in range(first_data_row, ws.max_row + 1):
        metric = ws.cell(row=r, column=1).value
        if not metric:
            continue
        if metric in values:
            continue
        row_vals = []
        for c in range(5, 5 + n_periods):
            v = ws.cell(row=r, column=c).value
            row_vals.append(v if is_num(v) else None)
        values[metric] = row_vals
        source_rows[metric] = r

    return values, periods, source_rows


def recalculate_with_excel(path):
    """Open `path` in a real Excel instance, force a full formula recalculation,
    and save the result to a new temp .xlsx file (the original file is left
    untouched - closed without saving). Returns the temp file's path.

    This is the automated equivalent of "open the export in Excel, let it
    recalculate, save it" - which is what surfaces the true formula-computed
    values instead of whatever was cached in the file at export time.
    """
    import win32com.client

    src_path = os.path.abspath(path)
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    os.remove(temp_path)  # SaveAs requires the path not exist yet on some Excel versions

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(src_path, UpdateLinks=0, ReadOnly=False)
        try:
            excel.CalculateFullRebuild()
            wb.SaveAs(temp_path, FileFormat=51)  # 51 = xlOpenXMLWorkbook (.xlsx)
        finally:
            wb.Close(SaveChanges=False)
    finally:
        excel.Quit()

    return temp_path


def g(values, name, i, default_zero_if_missing=False):
    """Get value of `name` at period index i. Returns None if missing/unavailable
    unless default_zero_if_missing is True (used for the debt-ratio silent-zero
    convention observed in the export)."""
    row = values.get(name)
    if row is None or i >= len(row):
        return 0.0 if default_zero_if_missing else None
    v = row[i]
    if v is None:
        return 0.0 if default_zero_if_missing else None
    return v


# ---------------------------------------------------------------------------
# Formula registry: metric_name -> function(values, i) -> (expected_value, note)
# `note` is None normally, or a short tag like "SILENT-ZERO-INPUT" when a
# missing input was substituted with 0 per the observed convention.
# ---------------------------------------------------------------------------

def formulas():
    F = {}

    def simple(name, fn, needs, zero_ok=()):
        """Register a formula. `needs` = list of input metric names (normal,
        missing => skip). `zero_ok` = subset of `needs` allowed to default to 0
        (debt-ratio silent-zero convention) instead of causing a skip."""
        def wrapped(values, i):
            inputs = {}
            note = None
            for n in needs:
                dz = n in zero_ok
                v = g(values, n, i, default_zero_if_missing=dz)
                if v is None:
                    return None, None  # can't verify - insufficient data
                if dz and (values.get(n) is None or values[n][i] is None):
                    note = "SILENT-ZERO-INPUT"
                inputs[n] = v
            try:
                return fn(inputs), note
            except ZeroDivisionError:
                return None, None
        F[name] = wrapped

    # --- Activity ratios ---
    simple("Cash Conversion Cycle", lambda x: x["DIO"] - x["DPO"] + x["DSO"],
           needs=["Days Inventory on Hand (DIO)", "Days Payables Outstanding (DPO)",
                  "Days Sales Outstanding (DSO)"])
    # (rename keys via a small local remap since dict keys must match `needs` names)
    F["Cash Conversion Cycle"] = lambda values, i: (
        (lambda dio, dpo, dso: (dio - dpo + dso, None) if None not in (dio, dpo, dso) else (None, None))(
            g(values, "Days Inventory on Hand (DIO)", i),
            g(values, "Days Payables Outstanding (DPO)", i),
            g(values, "Days Sales Outstanding (DSO)", i),
        )
    )
    F["Days Inventory on Hand (DIO)"] = lambda values, i: (
        (lambda t: (365.0 / t, None) if t not in (None, 0) else (None, None))(
            g(values, "Inventory Turnover", i))
    )
    F["Days Payables Outstanding (DPO)"] = lambda values, i: (
        (lambda t: (365.0 / t, None) if t not in (None, 0) else (None, None))(
            g(values, "Payables Turnover", i))
    )
    F["Days Sales Outstanding (DSO)"] = lambda values, i: (
        (lambda t: (365.0 / t, None) if t not in (None, 0) else (None, None))(
            g(values, "Receivables Turnover", i))
    )
    F["Inventory Turnover"] = lambda values, i: (
        (lambda cogs, avg: (abs(cogs) / avg, None) if None not in (cogs, avg) and avg != 0 else (None, None))(
            g(values, "Total Cost of Goods Sold (COGS)", i), g(values, "Average Inventory", i))
    )
    F["Payables Turnover"] = lambda values, i: (
        (lambda cogs, avg: (abs(cogs) / avg, None) if None not in (cogs, avg) and avg != 0 else (None, None))(
            g(values, "Total Cost of Goods Sold (COGS)", i), g(values, "Average Trade Payables", i))
    )
    F["Receivables Turnover"] = lambda values, i: (
        (lambda rev, avg: (rev / avg, None) if None not in (rev, avg) and avg != 0 else (None, None))(
            g(values, "Total Revenue", i), g(values, "Average Trade Receivables", i))
    )
    F["Total Assets Turnover"] = lambda values, i: (
        (lambda rev, avg: (rev / avg, None) if None not in (rev, avg) and avg != 0 else (None, None))(
            g(values, "Total Revenue", i), g(values, "Average Total Assets", i))
    )
    F["Working Capital Turnover"] = lambda values, i: (
        (lambda rev, avg: (rev / avg, None) if None not in (rev, avg) and avg != 0 else (None, None))(
            g(values, "Total Revenue", i), g(values, "Average Working Capital", i))
    )

    # --- Averages / rollups (pass through "Current"/"Previous" as reported) ---
    def avg_formula(current_name, previous_name):
        def fn(values, i):
            c = g(values, current_name, i)
            p = g(values, previous_name, i)
            if c is None or p is None:
                return None, None
            return (c + p) / 2.0, None
        return fn

    F["Average Inventory"] = avg_formula("Current Inventory", "Previous Inventory")
    F["Average Trade Payables"] = avg_formula("Current Trade Payables", "Previous Trade Payables")
    F["Average Trade Receivables"] = avg_formula("Current Trade Receivables", "Previous Trade Receivables")
    F["Average Total Assets"] = avg_formula("Current Total Assets", "Previous Total Assets")
    F["Average Working Capital"] = avg_formula("Current Working Capital", "Previous Working Capital")
    F["Average Total Shareholders' Equity"] = avg_formula(
        "Current Total Shareholders' Equity", "Previous Total Shareholders' Equity")

    F["Current Working Capital"] = lambda values, i: (
        (lambda a, l: (a - l, None) if None not in (a, l) else (None, None))(
            g(values, "Total Current Assets", i), g(values, "Total Current Liabilities", i))
    )
    F["Working Capital"] = F["Current Working Capital"]

    # --- Asset Valuation ---
    F["Book Value of Equity"] = lambda values, i: (
        (lambda a, l: (a - l, None) if None not in (a, l) else (None, None))(
            g(values, "Total Assets", i), g(values, "Total Liabilities", i))
    )

    # --- Coverage ---
    def cfo_finance_cost_coverage(values, i):
        cfo = g(values, "Net Cash from Operating Activities (CFO)", i)
        fc = g(values, "Finance Cost", i)
        zakat = g(values, "Income Tax / Zakat Expense", i)
        if None in (cfo, fc, zakat) or fc == 0:
            return None, None
        return (cfo - fc - zakat) / (-fc), None
    F["CFO Finance Cost Coverage"] = cfo_finance_cost_coverage

    F["Debt Coverage"] = lambda values, i: (
        (lambda cfo, debt: (cfo / debt, None) if None not in (cfo, debt) and debt != 0 else (None, None))(
            g(values, "Net Cash from Operating Activities (CFO)", i), g(values, "Total Debt", i))
    )
    F["Dividend Payment Coverage"] = lambda values, i: (
        (lambda cfo, div: (cfo / abs(div), None) if None not in (cfo, div) and div != 0 else (None, None))(
            g(values, "Net Cash from Operating Activities (CFO)", i), g(values, "CFF - Dividends Paid", i))
    )
    F["EBITDA Coverage"] = lambda values, i: (
        (lambda ebitda, fc: (ebitda / abs(fc), None) if None not in (ebitda, fc) and fc != 0 else (None, None))(
            g(values, "EBITDA", i), g(values, "Finance Cost", i))
    )
    F["Interest Coverage"] = lambda values, i: (
        (lambda ebit, fc: (ebit / abs(fc), None) if None not in (ebit, fc) and fc != 0 else (None, None))(
            g(values, "EBIT (Operating Income)", i), g(values, "Finance Cost", i))
    )
    F["OCF Debt Service Ratio (OCF DSR)"] = lambda values, i: (
        (lambda cfo, tds: (cfo / tds, None) if None not in (cfo, tds) and tds != 0 else (None, None))(
            g(values, "Net Cash from Operating Activities (CFO)", i), g(values, "Total Debt Service", i))
    )
    F["Operating Cash Flow to Liabilities"] = lambda values, i: (
        (lambda cfo, tl: (cfo / tl, None) if None not in (cfo, tl) and tl != 0 else (None, None))(
            g(values, "Net Cash from Operating Activities (CFO)", i), g(values, "Total Liabilities", i))
    )
    F["Reinvestment Ratio"] = lambda values, i: (
        (lambda cfo, capex: (cfo / abs(capex), None) if None not in (cfo, capex) and capex != 0 else (None, None))(
            g(values, "Net Cash from Operating Activities (CFO)", i), g(values, "CFI - Capital Expenditure (CapEx)", i))
    )

    # --- DuPont ---
    F["Interest Burden"] = lambda values, i: (
        (lambda ebt, ebit: (ebt / ebit, None) if None not in (ebt, ebit) and ebit != 0 else (None, None))(
            g(values, "EBT (Earnings Before Tax)", i), g(values, "EBIT (Operating Income)", i))
    )
    F["Tax Burden"] = lambda values, i: (
        (lambda np_, ebt: (np_ / ebt, None) if None not in (np_, ebt) and ebt != 0 else (None, None))(
            g(values, "Net Profit for the Period", i), g(values, "EBT (Earnings Before Tax)", i))
    )

    # --- Total Debt Service (sum of absolute financing outflows) ---
    def total_debt_service(values, i):
        parts = ["CFF - Finance Costs Paid", "CFF - Payment of Lease Liabilities",
                 "CFF - Repayment of Loans", "CFO - Finance Costs Paid"]
        total = 0.0
        any_present = False
        for p in parts:
            v = g(values, p, i)
            if v is not None:
                total += abs(v)
                any_present = True
        if not any_present:
            return None, None
        return total, None
    F["Total Debt Service"] = total_debt_service

    # --- Free Cash Flow family ---
    def fcf(values, i):
        cfo = g(values, "Net Cash from Operating Activities (CFO)", i)
        capex = g(values, "CFI - Capital Expenditure (CapEx)", i, default_zero_if_missing=True)
        if cfo is None:
            return None, None
        note = "SILENT-ZERO-INPUT" if values.get("CFI - Capital Expenditure (CapEx)") is None or \
            values["CFI - Capital Expenditure (CapEx)"][i] is None else None
        return cfo + capex, note
    F["Free Cash Flow (FCF)"] = fcf

    def fcfe(values, i):
        base, note = fcf(values, i)
        if base is None:
            return None, None
        nb = g(values, "Net Borrowing", i, default_zero_if_missing=True)
        return base + nb, note
    F["Free Cash Flow to Equity (FCFE)"] = fcfe

    def fcff(values, i):
        cfo = g(values, "Net Cash from Operating Activities (CFO)", i)
        capex = g(values, "CFI - Capital Expenditure (CapEx)", i, default_zero_if_missing=True)
        fc = g(values, "Finance Cost", i, default_zero_if_missing=True)
        if cfo is None:
            return None, None
        return cfo + capex + abs(fc), None
    F["Free Cash Flow to Firm - FCFF (Tax Rate Assumed Zero)"] = fcff

    # --- Liquidity ---
    F["Cash Ratio"] = lambda values, i: (
        (lambda cash, tcl: (cash / tcl, None) if None not in (cash, tcl) and tcl != 0 else (None, None))(
            g(values, "Cash & Cash Equivalents", i), g(values, "Total Current Liabilities", i))
    )
    F["Current Ratio"] = lambda values, i: (
        (lambda tca, tcl: (tca / tcl, None) if None not in (tca, tcl) and tcl != 0 else (None, None))(
            g(values, "Total Current Assets", i), g(values, "Total Current Liabilities", i))
    )
    F["Quick Ratio"] = lambda values, i: (
        (lambda cash, ar, tcl: ((cash + ar) / tcl, None) if None not in (cash, ar, tcl) and tcl != 0 else (None, None))(
            g(values, "Cash & Cash Equivalents", i), g(values, "Trade Receivables", i),
            g(values, "Total Current Liabilities", i))
    )

    # --- Margins ---
    def margin(numerator_name):
        def fn(values, i):
            num = g(values, numerator_name, i)
            rev = g(values, "Total Revenue", i)
            if None in (num, rev) or rev == 0:
                return None, None
            return num / rev, None
        return fn
    F["EBIT Margin"] = margin("EBIT (Operating Income)")
    F["EBITDA Margin"] = margin("EBITDA")
    F["Gross Profit Margin"] = margin("Gross Profit")
    F["Net Profit Margin"] = margin("Net Profit for the Period")
    F["Operating Profit Margin"] = margin("EBIT (Operating Income)")
    F["Pretax Profit Margin"] = margin("EBT (Earnings Before Tax)")

    # --- Performance ---
    def capex_to_dep(values, i):
        capex = g(values, "CFI - Capital Expenditure (CapEx)", i, default_zero_if_missing=True)
        dep = g(values, "CFO - Depreciation Adjustment", i)
        if dep is None or dep == 0:
            return None, None
        return abs(capex) / dep, None
    F["CapEx to Depreciation"] = capex_to_dep

    def capex_to_revenue(values, i):
        capex = g(values, "CFI - Capital Expenditure (CapEx)", i, default_zero_if_missing=True)
        rev = g(values, "Total Revenue", i)
        if rev is None or rev == 0:
            return None, None
        return abs(capex) / rev, None
    F["CapEx to Revenue"] = capex_to_revenue

    F["Cash Flow Quality"] = lambda values, i: (
        (lambda cfo, npf: (cfo / npf, None) if None not in (cfo, npf) and npf != 0 else (None, None))(
            g(values, "Net Cash from Operating Activities (CFO)", i), g(values, "Net Profit for the Period", i))
    )
    F["Cash Flow to Revenue"] = lambda values, i: (
        (lambda cfo, rev: (cfo / rev, None) if None not in (cfo, rev) and rev != 0 else (None, None))(
            g(values, "Net Cash from Operating Activities (CFO)", i), g(values, "Total Revenue", i))
    )
    F["Cash Return on Assets"] = lambda values, i: (
        (lambda cfo, avg: (cfo / avg, None) if None not in (cfo, avg) and avg != 0 else (None, None))(
            g(values, "Net Cash from Operating Activities (CFO)", i), g(values, "Average Total Assets", i))
    )
    F["Cash Return on Equity"] = lambda values, i: (
        (lambda cfo, avg: (cfo / avg, None) if None not in (cfo, avg) and avg != 0 else (None, None))(
            g(values, "Net Cash from Operating Activities (CFO)", i),
            g(values, "Average Total Shareholders' Equity", i))
    )
    F["Cash to Operating Income"] = lambda values, i: (
        (lambda cfo, ebit: (cfo / ebit, None) if None not in (cfo, ebit) and ebit != 0 else (None, None))(
            g(values, "Net Cash from Operating Activities (CFO)", i), g(values, "EBIT (Operating Income)", i))
    )
    F["Dividend Payout Ratio"] = lambda values, i: (
        (lambda div, npf: (abs(div) / npf, None) if None not in (div, npf) and npf != 0 else (None, None))(
            g(values, "CFF - Dividends Paid", i), g(values, "Net Profit for the Period", i))
    )

    def retention_ratio(values, i):
        payout, note = F["Dividend Payout Ratio"](values, i)
        if payout is None:
            return None, None
        return 1 - payout, note
    F["Retention Ratio"] = retention_ratio

    # --- Profitability ---
    def ebitda(values, i):
        ebit = g(values, "EBIT (Operating Income)", i)
        dep = g(values, "CFO - Depreciation Adjustment", i, default_zero_if_missing=True)
        amort = g(values, "CFO - Amortization Adjustment", i, default_zero_if_missing=True)
        if ebit is None:
            return None, None
        return ebit + dep + amort, None
    F["EBITDA"] = ebitda

    F["Gross Profit"] = lambda values, i: (
        (lambda rev, cogs: (rev + cogs, None) if None not in (rev, cogs) else (None, None))(
            g(values, "Total Revenue", i), g(values, "Total Cost of Goods Sold (COGS)", i))
    )
    F["NOPAT (Tax Rate Assumed Zero)"] = lambda values, i: (
        (lambda ebit: (ebit, None) if ebit is not None else (None, None))(
            g(values, "EBIT (Operating Income)", i))
    )

    # --- Profitability Ratios ---
    F["Operating Return on Assets"] = lambda values, i: (
        (lambda ebit, avg: (ebit / avg, None) if None not in (ebit, avg) and avg != 0 else (None, None))(
            g(values, "EBIT (Operating Income)", i), g(values, "Average Total Assets", i))
    )

    def roa_adjusted(values, i):
        npf = g(values, "Net Profit for the Period", i)
        fc = g(values, "Finance Cost", i, default_zero_if_missing=True)
        avg = g(values, "Average Total Assets", i)
        if None in (npf, avg) or avg == 0:
            return None, None
        return (npf + abs(fc)) / avg, None
    F["ROA Adjusted (Tax Rate Assumed Zero)"] = roa_adjusted

    def roe_dupont3(values, i):
        npm, _ = F["Net Profit Margin"](values, i)
        tat, _ = F["Total Assets Turnover"](values, i)
        fl, _ = F["Financial Leverage"](values, i)
        if None in (npm, tat, fl):
            return None, None
        return npm * tat * fl, None
    F["ROE (DuPont 3-Factor)"] = roe_dupont3

    def roe_dupont5(values, i):
        tb, _ = F["Tax Burden"](values, i)
        ib, _ = F["Interest Burden"](values, i)
        em, _ = F["EBIT Margin"](values, i)
        tat, _ = F["Total Assets Turnover"](values, i)
        fl, _ = F["Financial Leverage"](values, i)
        if None in (tb, ib, em, tat, fl):
            return None, None
        return tb * ib * em * tat * fl, None
    F["ROE (DuPont 5-Factor)"] = roe_dupont5

    F["Return on Assets (ROA)"] = lambda values, i: (
        (lambda npf, avg: (npf / avg, None) if None not in (npf, avg) and avg != 0 else (None, None))(
            g(values, "Net Profit for the Period", i), g(values, "Average Total Assets", i))
    )
    F["Return on Equity (ROE)"] = lambda values, i: (
        (lambda npf, avg: (npf / avg, None) if None not in (npf, avg) and avg != 0 else (None, None))(
            g(values, "Net Profit for the Period", i), g(values, "Average Total Shareholders' Equity", i))
    )

    # --- Solvency ---
    DEBT_ZERO_OK = {"Total Debt", "Non Current Debt", "Current Debt"}

    def debt_to_assets(values, i):
        debt = g(values, "Total Debt", i, default_zero_if_missing=True)
        assets = g(values, "Total Assets", i)
        if assets is None or assets == 0:
            return None, None
        note = "SILENT-ZERO-INPUT" if values.get("Total Debt") is None or values["Total Debt"][i] is None else None
        return debt / assets, note
    F["Debt to Assets"] = debt_to_assets

    def debt_to_capital(values, i):
        debt = g(values, "Total Debt", i, default_zero_if_missing=True)
        equity = g(values, "Total Shareholders' Equity", i)
        if equity is None or (debt + equity) == 0:
            return None, None
        note = "SILENT-ZERO-INPUT" if values.get("Total Debt") is None or values["Total Debt"][i] is None else None
        return debt / (debt + equity), note
    F["Debt to Capital"] = debt_to_capital

    def debt_to_ebitda(values, i):
        debt = g(values, "Total Debt", i, default_zero_if_missing=True)
        ebitda_v = g(values, "EBITDA", i)
        if ebitda_v is None or ebitda_v == 0:
            return None, None
        note = "SILENT-ZERO-INPUT" if values.get("Total Debt") is None or values["Total Debt"][i] is None else None
        return debt / ebitda_v, note
    F["Debt to EBITDA"] = debt_to_ebitda

    def debt_to_equity(values, i):
        debt = g(values, "Total Debt", i, default_zero_if_missing=True)
        equity = g(values, "Total Shareholders' Equity", i)
        if equity is None or equity == 0:
            return None, None
        note = "SILENT-ZERO-INPUT" if values.get("Total Debt") is None or values["Total Debt"][i] is None else None
        return debt / equity, note
    F["Debt to Equity"] = debt_to_equity

    F["Financial Leverage"] = lambda values, i: (
        (lambda avga, avge: (avga / avge, None) if None not in (avga, avge) and avge != 0 else (None, None))(
            g(values, "Average Total Assets", i), g(values, "Average Total Shareholders' Equity", i))
    )

    def net_debt(values, i):
        debt = g(values, "Total Debt", i, default_zero_if_missing=True)
        cash = g(values, "Cash & Cash Equivalents", i)
        if cash is None:
            return None, None
        note = "SILENT-ZERO-INPUT" if values.get("Total Debt") is None or values["Total Debt"][i] is None else None
        return debt - cash, note
    F["Net Debt"] = net_debt

    def net_debt_to_ebitda(values, i):
        nd, note = net_debt(values, i)
        ebitda_v = g(values, "EBITDA", i)
        if nd is None or ebitda_v is None or ebitda_v == 0:
            return None, None
        return nd / ebitda_v, note
    F["Net Debt to EBITDA"] = net_debt_to_ebitda

    F["Total Debt"] = lambda values, i: (
        (lambda cd, ncd: (cd + ncd, None) if None not in (cd, ncd) else (None, None))(
            g(values, "Current Debt", i), g(values, "Non Current Debt", i))
    )

    return F


def check_values(xlsx_path, values, periods, source_rows, tolerance, csv_path, fail_only, label):
    """Run the recompute-and-compare checks for one (values, periods, source_rows)
    dataset (as returned by load_metric_values), print a report, and optionally
    write a CSV. `label` is just a header string to distinguish this run's
    output (e.g. "WEB (cached export values)" vs "EXCEL (recalculated)")."""
    F = formulas()

    results = []  # (metric, period, status, reported, expected, diff_pct, note, source_row)
    for metric, fn in F.items():
        reported_row = values.get(metric)
        if reported_row is None:
            continue
        source_row = source_rows.get(metric)
        for i, period in enumerate(periods):
            reported = reported_row[i] if i < len(reported_row) else None
            expected, note = fn(values, i)
            if expected is None:
                continue  # insufficient data to verify - not a failure
            if reported is None:
                results.append((metric, period, "REPORTED-MISSING", reported, expected, None, note, source_row))
                continue
            if abs(reported) < ABS_EPS and abs(expected) < ABS_EPS:
                status = "PASS"
                diff_pct = 0.0
            elif abs(reported) < ABS_EPS:
                status = "FAIL"
                diff_pct = None
            else:
                diff_pct = abs(expected - reported) / abs(reported)
                status = "PASS" if diff_pct <= tolerance else "FAIL"
            results.append((metric, period, status, reported, expected, diff_pct, note, source_row))

    fails = [r for r in results if r[2] == "FAIL"]
    silent_zero = [r for r in results if r[6] == "SILENT-ZERO-INPUT" and r[2] == "PASS"]
    reported_missing = [r for r in results if r[2] == "REPORTED-MISSING"]
    total_checked = len(results)

    # --- Diagnostic: for FAILs whose formula depends on an "Average X" input,
    # check whether the reported value instead matches a variant using the
    # CURRENT (period-end) balance in place of the average. This doesn't
    # change the FAIL verdict - it just helps triage the likely cause. ---
    AVG_TO_CURRENT = {
        "Average Total Assets": "Current Total Assets",
        "Average Inventory": "Current Inventory",
        "Average Trade Payables": "Current Trade Payables",
        "Average Trade Receivables": "Current Trade Receivables",
        "Average Working Capital": "Current Working Capital",
        "Average Total Shareholders' Equity": "Current Total Shareholders' Equity",
    }
    RATIO_AVG_INPUT = {
        "Inventory Turnover": ("Total Cost of Goods Sold (COGS)", "Average Inventory", "abs_div"),
        "Payables Turnover": ("Total Cost of Goods Sold (COGS)", "Average Trade Payables", "abs_div"),
        "Receivables Turnover": ("Total Revenue", "Average Trade Receivables", "div"),
        "Total Assets Turnover": ("Total Revenue", "Average Total Assets", "div"),
        "Working Capital Turnover": ("Total Revenue", "Average Working Capital", "div"),
        "Cash Return on Assets": ("Net Cash from Operating Activities (CFO)", "Average Total Assets", "div"),
        "Cash Return on Equity": ("Net Cash from Operating Activities (CFO)", "Average Total Shareholders' Equity", "div"),
        "Operating Return on Assets": ("EBIT (Operating Income)", "Average Total Assets", "div"),
        "ROA Adjusted (Tax Rate Assumed Zero)": (None, "Average Total Assets", "roa_adj"),
        "Return on Assets (ROA)": ("Net Profit for the Period", "Average Total Assets", "div"),
        "Return on Equity (ROE)": ("Net Profit for the Period", "Average Total Shareholders' Equity", "div"),
        "Financial Leverage": ("Average Total Assets", "Average Total Shareholders' Equity", "leverage"),
    }

    diagnosed = {}
    for metric, period, status, reported, expected, diff_pct, note, source_row in fails:
        i = periods.index(period)
        spec = RATIO_AVG_INPUT.get(metric)
        if not spec:
            continue
        num_name, avg_name, mode = spec
        current_name = AVG_TO_CURRENT.get(avg_name)
        cur_val = g(values, current_name, i) if current_name else None
        alt = None
        if mode == "div" and num_name:
            num = g(values, num_name, i)
            if num is not None and cur_val not in (None, 0):
                alt = num / cur_val
        elif mode == "abs_div" and num_name:
            num = g(values, num_name, i)
            if num is not None and cur_val not in (None, 0):
                alt = abs(num) / cur_val
        elif mode == "leverage":
            cur_a = g(values, "Current Total Assets", i)
            cur_e = g(values, "Current Total Shareholders' Equity", i)
            if cur_a is not None and cur_e not in (None, 0):
                alt = cur_a / cur_e
        elif mode == "roa_adj":
            npf = g(values, "Net Profit for the Period", i)
            fc = g(values, "Finance Cost", i, default_zero_if_missing=True)
            if npf is not None and cur_val not in (None, 0):
                alt = (npf + abs(fc)) / cur_val
        if alt is not None and reported not in (None, 0):
            matches_current_variant = abs(alt - reported) / abs(reported) <= tolerance
            diagnosed[(metric, period)] = matches_current_variant

    print(f"Lameh Ratio Verification Report [{label}]")
    print(f"File: {xlsx_path}")
    print(f"Periods: {periods}")
    print(f"Tolerance: {tolerance:.2%}")
    print("=" * 70)
    print(f"Total (metric, period) checks performed: {total_checked}")
    print(f"  PASS: {len(results) - len(fails) - len(reported_missing)}")
    print(f"  FAIL: {len(fails)}")
    print(f"  Reported value missing but computable: {len(reported_missing)}")
    print(f"  Passed using a silently-zeroed missing input: {len(silent_zero)}")
    print("=" * 70)

    if fails:
        print("\nFAILURES (recomputed value does not match reported value):\n")
        for metric, period, status, reported, expected, diff_pct, note, source_row in fails:
            dp = f"{diff_pct:.2%}" if diff_pct is not None else "n/a (reported=0)"
            print(f"  [{metric}] period={period} (Excel row {source_row})")
            print(f"      reported = {reported}")
            print(f"      expected (using the properly-averaged/documented formula) = {expected}")
            print(f"      diff     = {dp}")
            diag = diagnosed.get((metric, period))
            if diag is True:
                print(f"      diagnosis = MATCHES a period-end-balance variant instead of the average")
            elif diag is False:
                print(f"      diagnosis = does not match the average formula OR a simple period-end-balance variant")
            if note:
                print(f"      note     = {note}")
            print()
    else:
        print("\nNo failures found within tolerance.")

    if reported_missing:
        print("\nCOMPUTABLE BUT NOT REPORTED (Lameh returned no value, but inputs exist):\n")
        for metric, period, status, reported, expected, diff_pct, note, source_row in reported_missing:
            print(f"  [{metric}] period={period} (Excel row {source_row}) -> could compute {expected}, but Lameh shows no value")

    if silent_zero:
        print(f"\nNOTE: {len(silent_zero)} checks passed only because a missing debt-related "
              f"input was treated as 0 (see docstring). These are not failures, but a company "
              f"with missing data currently looks identical to a company with zero debt.")
        metrics_affected = sorted(set(r[0] for r in silent_zero))
        print(f"  Affected ratios: {', '.join(metrics_affected)}")

    if csv_path:
        import csv
        rows_to_write = [r for r in results if r[2] != "PASS"] if fail_only else results
        parent_dir = os.path.dirname(csv_path)
        if parent_dir:
            os.makedirs(parent_dir, exist_ok=True)
        with open(csv_path, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Metric", "Period", "Status", "Reported", "Expected", "Diff %", "Note", "Excel Row"])
            for row in rows_to_write:
                w.writerow(row)
        suffix = " (fail-only)" if fail_only else ""
        print(f"\nCSV report written to {csv_path}{suffix}")


def _with_suffix(path, tag):
    """Insert `tag` (e.g. "-web") right before the file extension."""
    base, ext = os.path.splitext(path)
    return f"{base}{tag}{ext}"


def main():
    # Company names can contain non-ASCII (e.g. Arabic) text, which crashes on
    # Windows consoles stuck on a legacy codepage (cp1252) unless stdout is
    # explicitly put into UTF-8 mode.
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    ap = argparse.ArgumentParser(description="Verify Lameh Sector Analysis ratio export.")
    ap.add_argument("xlsx_path")
    ap.add_argument("--tolerance", type=float, default=TOLERANCE_DEFAULT,
                     help="Relative tolerance (default 0.5%%)")
    ap.add_argument("--csv", default=None,
                     help="Optional base path to write CSV reports. Two files are written: "
                          "<name>-web<ext> (cached export values) and <name>-excel<ext> "
                          "(values after a real Excel recalculation). Defaults to "
                          f"{DEFAULT_CSV_DIR}/<date>/<company name>.")
    ap.add_argument("--fail-only", action="store_true",
                     help="When writing --csv, include only non-PASS rows (FAIL / REPORTED-MISSING)")
    ap.add_argument("--web-only", action="store_true",
                     help="Skip the Excel-recalculation pass (no Excel/pywin32 required)")
    args = ap.parse_args()

    csv_base = args.csv
    if csv_base is None:
        date = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        company = get_company_name(args.xlsx_path)
        filename = sanitize_filename(company) if company else "unknown-company"
        csv_base = os.path.join(DEFAULT_CSV_DIR, date, f"{filename}.csv")

    web_csv = _with_suffix(csv_base, "-web")
    excel_csv = _with_suffix(csv_base, "-excel")

    values_web, periods_web, source_rows_web = load_metric_values(args.xlsx_path)
    check_values(args.xlsx_path, values_web, periods_web, source_rows_web,
                 args.tolerance, web_csv, args.fail_only, label="WEB (cached export values)")

    if args.web_only:
        return

    print("\n" + "=" * 70)
    print("Recalculating formulas with Excel (this opens Excel in the background)...")
    try:
        recalculated_path = recalculate_with_excel(args.xlsx_path)
    except Exception as e:
        print(f"Could not recalculate with Excel ({e}). Skipping the EXCEL report.\n"
              f"(Requires pywin32 and a local Excel installation - use --web-only to skip this.)")
        return

    try:
        values_excel, periods_excel, source_rows_excel = load_metric_values(recalculated_path)

        print()
        check_values(args.xlsx_path, values_excel, periods_excel, source_rows_excel,
                     args.tolerance, excel_csv, args.fail_only, label="EXCEL (recalculated formulas)")
    finally:
        os.remove(recalculated_path)


if __name__ == "__main__":
    main()